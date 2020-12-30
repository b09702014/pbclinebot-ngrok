# 插入設置虛擬環境與line聊天機器人所需的套件
from django.shortcuts import render
from django.conf import settings
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt

from linebot import LineBotApi, WebhookParser
from linebot.exceptions import InvalidSignatureError, LineBotApiError
from linebot.models import *
line_bot_api = LineBotApi(settings.LINE_CHANNEL_ACCESS_TOKEN)
parser = WebhookParser(settings.LINE_CHANNEL_SECRET)
from 聊天機器人.models import *

# 插入主程式需要的套件
import datetime
import matplotlib
import matplotlib.pyplot as py
import matplotlib.dates as mdates
py.switch_backend('agg')
import pyimgur
import requests, lxml
from bs4 import BeautifulSoup
import re
import xlrd
import operator
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import time
import random

# 設置可以回傳訊息的聊天機器人
@csrf_exempt
def callback(request):
    if request.method == 'POST':  # 連結django與linechatbox並傳送訊息
        signature = request.META['HTTP_X_LINE_SIGNATURE']
        body = request.body.decode('utf-8')

        message=[]  # linebot訊息回傳list
        today = datetime.date.today()  # 紀錄當日日期

        # CLASS userinfo: 儲存關於使用者的所有資訊
        class userInfo:
            def __init__(self, name, age, gender, height, weight, mode, bmr, tdee):
                self.name = name  # 使用者姓名
                self.age = age  # 使用者年齡
                self.gender = gender  # 使用者性別
                self.height = height  # 使用者身高
                self.weight = weight  # 使用者體重
                self.mode = mode  # 使用者設定的策略模式
                self.bmr = bmr  # 使用者的基礎代謝率
                self.tdee = tdee  # 使用者每日可攝取的熱量

            def changeMode(self, mCode):
                # 改變使用者的策略模式
                self.mode = mode

            def printUser(self):
                # 印出所有使用者的資訊
                userStr = vars(self)
                for item in userStr:
                    print(item, ':', userStr[item])

        # CLASS: foody: 儲存關於食物的資訊
        class foody:
            def __init__(self, fDate, fType, food, cal):
                self.date = fDate  # 食物攝取日期
                self.type = fType  # 食物攝取類別
                self.food = food  # 食物
                self.cal = cal  # 食物熱量
        """
        主程式函數及爬蟲
        """
        def bmrCal(age, gender, height, weight):  # 計算使用者的BMR
            if gender == 'm':
                bmr = 66.47 + (13.75 * weight) + (5.003 * height) - (6.755 * age)
            elif gender == 'f':
                bmr = 447.593 + (9.247 * weight) + (3.098 * height) - (4.330 * age)
            return bmr

        def tdeeCal(bmr, habit):  # 計算使用者的TDEE
            if habit == 'A':
                tdee = bmr*1.2
            elif habit == 'B':
                tdee = bmr*1.375
            elif habit == 'C':
                tdee = bmr*1.55
            elif habit == 'D':
                tdee = bmr*1.725
            elif habit == 'E':
                tdee = bmr*1.9
            return tdee    

        calDict = {}  # 食物熱量對照字典
        typeDict = {'早餐': 0, '午餐': 0, '晚餐': 0, '宵夜': 0, '飲料': 0, '其他': 0}  # 用餐類別字典

        # 尋找食物熱量的方法一: 利用爬蟲讀取網站資料
        list_rows = []  # 爬蟲資料串列
        url = "http://www.dpjh.ylc.edu.tw/97/217health/4-4-01calorie.htm#1"
        resp = requests.get(url)
        resp.encoding = 'big5'

        soup = BeautifulSoup(resp.text, "lxml")
        rows = soup.find_all('tr')
        
        for row in rows:  # 爬取網站資料並清理
            row_td = row.find_all('td')
            str_cells = str(row_td) 
            clean = re.compile('<.*?>')
            clean2 = (re.sub(clean, '',str_cells))
            list_rows.append(clean2)
        
        for i in range(4,len(list_rows)):  #  將爬蟲資料整理後存到字典
            if list_rows[i].find(',') != -1:
                list_rows[i] = list_rows[i].strip('[]')
                list_rows[i] = list_rows[i].split(', ')
                if list_rows[i][0] != "種類":
                    try:
                        calDict[list_rows[i][0]] = int(list_rows[i][3])
                    except:
                        list_rows[i][3] = list_rows[i][3].strip("卡")  # 去掉儲存格內熱量的國字表示
                        calDict[list_rows[i][0]] = int(list_rows[i][3])

        # 尋找食物熱量的方法二: 讀取 excel檔案的食物熱量對照表
        readbook = xlrd.open_workbook(r'C:\Users\林煥傑\Desktop\2020食物熱量對照大全.xlsx')
        foodsheet = readbook.sheets()[0]
        nrows = foodsheet.nrows
        for i in range(nrows - 1):
            food = foodsheet.cell(i + 1,0).value
            cal = foodsheet.cell(i + 1,1).value
            calDict[food] = float(cal)
        
        calKeys = calDict.keys()  # 熱量字典的Keys
        calKeys_sorted = sorted(calKeys, key=len, reverse=True)  # 將熱量字度依長度排序
        max_len = len(calKeys_sorted[0])  # 查看熱量字典最長的字是多長
        # 對聊天機器人輸入指令
        try:
            events = parser.parse(body, signature)
        except InvalidSignatureError:
            return HttpResponseForbidden()
        except LineBotApiError:
            return HttpResponseBadRequest()

        for event in events:
            if isinstance(event, MessageEvent):
                mtext=event.message.text                # 使用者傳的訊息
                uid=event.source.user_id                # 使用者的UID(每個人不同)
                profile=line_bot_api.get_profile(uid)
                name=profile.display_name               # 使用者的line稱呼

            #往下為功能列表，分為：
            #1.建立與修改個人資料
            #2.查詢資料：輸出當日可攝取熱量及當日熱量圓餅圖
            #3.輸入食物：此處為食物有在字典內或是切字有食物的情況
            #4.手動輸入食物：字典及切字皆找不到時可自行輸入，輸入食物及熱量會增加到2020食物熱量對照大全.xlsx中
            #5.每月報告：輸入關鍵字可回傳當日往前31天飲食熱量折線圖
            #6.策略更改：增重、正常、減重提醒
            #7.使用說明
            #其中2、5、7功能有放在圖表選單，其他皆需要手動輸入


                if ' ' in mtext:  # 1.BMR計算(性別, 年齡, 身高, 體重, 運動習慣)以及個人資料紀錄或修改
                    try:
                        data = mtext.split(" ")
                        age = int(data[1])  # 使用者年齡
                        gender = data[0]  # 使用者性別
                        height = float(data[2])  # 使用者身高
                        weight = float(data[3])  # 使用者體重
                        habit = data[4]  # 使用者的運動習慣
                        mode = 'n'  # 預設使用者的策略為正常

                        bmr = bmrCal(age=age, gender=gender, height=height, weight=weight)  # 計算使用者的BMR
                        tdee = tdeeCal(bmr, habit)  # 計算使用者的TDEE
  
                        # user = user_Info(name=name, age=age, gender=gender, height=height, weight=weight,mode=mode, bmr=bmr, tdee=tdee)

                        # 印出使用者的基礎代謝率和TDEE
                        message.append(TextSendMessage(text='你的每日基礎代謝率(BMR)是' + str(round(bmr, 3)) + ' Kcal/Day。'))
                        message.append(TextSendMessage(text='你每日可攝取的熱量是' + str(round(tdee, 3)) + ' Kcal/Day。'))

                        # 儲存使用者的所有資訊
                        if User_Info.objects.filter(uid=uid,data_type="個人資料",number=1).exists() == False:
                            User_Info.objects.create(uid=uid,name=name,data_type="個人資料",mtext=mtext,user_bmr=str(tdee), cal=str(tdee), number=1)
                            message.append(TextSendMessage(text='個人資料新增完畢'))
                        else:
                            User_Info.objects.filter(uid=uid,data_type="個人資料",number=1).update(user_bmr=str(tdee),cal=str(tdee))
                            User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).update(user_bmr=str(tdee),cal=str(tdee))
                            User_Info.objects.filter(uid=uid, data_type="吃東西", date=today).delete()
                            message.append(TextSendMessage(text='已修改個人資料'))
                    except:  # 除錯(使用者輸入錯誤提醒)
                        message.append(TextSendMessage(text="輸入格式為：性別 年齡 身高 體重 運動習慣"))
                

                elif "查詢資料" in mtext:  # 2.查詢今天的熱量，輸出圓餅圖

                    """
                    第一部分：查詢當日熱量
                    """
                    # 先確認使用者有無輸入基本資料
                    if User_Info.objects.filter(uid=uid,data_type="個人資料",number=1).exists() == True:
                        infor = User_Info.objects.get(uid=uid,data_type="個人資料",number=1)

                        # 替使用者建立當日資料
                        if User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).exists() == False:
                            User_Info.objects.create(uid=uid,data_type="個人資料",name=name,user_bmr=infor.user_bmr,cal=infor.user_bmr)

                        # 當天的使用者紀錄
                        user = User_Info.objects.get(uid=uid,data_type="個人資料",date=today)
 
                        # 回傳使用者熱量或是提醒
                        if float(user.cal) > 0:
                            message.append(TextSendMessage(text='今天仍可以攝取' + str(round(float(user.cal), 3))))
                        else:
                            message.append(StickerSendMessage(package_id=1, sticker_id=118))
                            message.append(TextSendMessage(text="你吃太多了！該休息了喔！"))

                        """
                        第二部分：回傳當日圓餅圖
                        """
                        # 資料庫抓取使用者當日飲食紀錄
                        meals = User_Info.objects.filter(uid=uid, data_type="吃東西", date=today)
                        if len(meals) != 0:  # 有飲食紀錄才製作圖表
                            for meal in meals:
                                typeDict[meal.period] += float(meal.cal)  # 計算每個種類的總熱量

                            typeDict_keys = ['Breakfast', 'Lunch', 'Dinner', 'Snacks', 'Drinks', 'Others']
                            typeDict_values = list(typeDict.values())
                            for data in typeDict_values:
                                data = float(data)
                            
                            #補充文字(比較)
                            eat_str = '今天吃了：  '
                            periodDict = dict()
                            for period in typeDict:
                                if typeDict[period] != 0.0:
                                    eat_str += period + ":" + str(typeDict[period]) +"大卡 "
                            
                            message.append(TextSendMessage(text=eat_str))

                            labels = []
                            dataSet = []
                            for i in range(6):
                                if typeDict_values[i] != 0:
                                    labels.append(typeDict_keys[i])
                                    dataSet.append(typeDict_values[i])
                            
                            py.clf()
                            py.pie(dataSet, labels=labels, autopct='%1.1f%%', startangle=90)
                            py.axis('equal')
                            py.tight_layout()
                            py.title('Yours Calorie Intake Percentage Today')
                            pieName = str(random.random())
                            py.savefig('C:\\Users\\林煥傑\\Desktop\\' + pieName + '.png')
                            py.legend(loc = "center right")

                            # 取得圓餅圖URL
                            CLIENT_ID = "45c5f1f50f87a64"
                            PATH = 'C:\\Users\\林煥傑\\Desktop\\' + pieName + '.png'
                            im = pyimgur.Imgur(CLIENT_ID)
                            uploaded_image = im.upload_image(PATH, title="Uploaded with PyImgur")
                            # linebot傳送圖表
                            message.append(ImageSendMessage(original_content_url=uploaded_image.link, preview_image_url=uploaded_image.link))

                    else:  # 使用者尚未輸入資料提醒(原則上會在一開始提醒使用者先輸入基本資料)
                        message.append(TextSendMessage(text="你還沒輸入資料喔"))

                elif "/" in mtext:  # 3.輸入每天吃了什麼
                    calHere = 0
                    try:
                        ftype, food = mtext.split("/")
                        if ftype in typeDict:
                            if food in calDict:  # 若可以在CalDict中找到食物
                                # 建立飲食資料
                                User_Info.objects.create(uid=uid,name=name,mtext=mtext,data_type="吃東西",food=food,period=ftype,cal=calDict[food])

                                # 先確認使用者有無輸入基本資料
                                if User_Info.objects.filter(uid=uid,data_type="個人資料",number=1).exists() == True:
                                    infor = User_Info.objects.get(uid=uid,data_type="個人資料",number=1)  # number指的是原始資料

                                    # 替使用者建立當日資料
                                    if User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).exists() == False:
                                        User_Info.objects.create(uid=uid,data_type="個人資料",name=name,user_bmr=infor.user_bmr,cal=infor.user_bmr)
                            
                                    user = User_Info.objects.get(uid=uid,data_type="個人資料",date=today)  # 當天的使用者紀錄
                                    calculate = float(user.cal) - float(calDict[food])  # 計算當日熱量
                                    User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).update(cal=str(calculate))

                                    # linebot回傳訊息
                                    message.append(TextSendMessage(text=food + '的熱量是' + str(calDict[food]) + 'Kcal'))
                                    message.append(TextSendMessage(text='今天仍可以攝取' + str(round(calculate, 3)) + 'Kcal'))
                                else:
                                    message.append(TextSendMessage(text="你還沒輸入個人資料喔"))
                            else:
                                foodLen = len(food)
                                # 切割文字，查詢是否有類似的食物
                                if foodLen >= max_len:
                                    maxStr = max_len
                                else:
                                    maxStr = foodLen
                                totalBreak = False
                                while totalBreak is False:
                                    for j in range(len(food)):
                                        for i in range(maxStr, 0, -1):
                                            item = False
                                            if (i+j) > len(food):
                                                continue
                                            else:
                                                check_str = food[j:i+j]
                                                for fd in calKeys_sorted:
                                                    if check_str == fd:
                                                        calHere += calDict[fd]
                                                        food = food[(1+j):]
                                                        item = True
                                                        break
                                            if item is True:
                                                break
                                        if item is True:
                                            break
                                    totalBreak = True
                                    for word in calKeys_sorted:
                                        if word in food:
                                            totalBreak = False
                                            break
                                if calHere == 0:
                                    message.append(TextSendMessage(text="抱歉,我們沒有您輸入食物的熱量喔。請使用輸入熱量功能。"))
                                else:
                                    # 建立飲食資料
                                    User_Info.objects.create(uid=uid,name=name,mtext=mtext,data_type="吃東西",food=food,period=ftype,cal=calHere)

                                    # 從基本資料更新當日可攝取熱量
                                    infor = User_Info.objects.get(uid=uid,data_type="個人資料",number=1)

                                    # 替使用者建立當日資料
                                    if User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).exists() == False:
                                        User_Info.objects.create(uid=uid,data_type="個人資料",name=name,user_bmr=infor.user_bmr,cal=infor.user_bmr)

                                    user = User_Info.objects.get(uid=uid,data_type="個人資料",date=today)  # 當天的使用者紀錄
                                    calculate = float(user.cal) - float(calHere)  # 計算當日熱量
                                    User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).update(cal=str(calculate))

                                    # linebot回傳訊息
                                    message.append(TextSendMessage(text=food + '預估的熱量是' + str(calHere) + 'Kcal'))
                                    message.append(TextSendMessage(text='今天仍可以攝取' + str(round(calculate, 3)) + 'Kcal'))
                        else:  # 輸入格式錯誤提醒
                            message.append(TextSendMessage(text="飲食輸入格式為：時段(早餐、午餐、晚餐、宵夜、飲料、其他)/食物，"))
                    except:  # 輸入格式錯誤提醒
                        message.append(TextSendMessage(text="飲食輸入格式為：時段(早餐、午餐、晚餐、宵夜、飲料、其他)/食物，"))

                elif "，" in mtext:  # 4.自行輸入食物熱量(時段，食物，熱量)
                    period, food, cal = mtext.split("，")  # 此處設定為全形逗號(中文輸入)

                    # 建立飲食資料
                    User_Info.objects.create(uid=uid,name=name,mtext=mtext,data_type="吃東西",food=food,period=period,cal=cal)

                    # 將食物及熱量加入2020食物熱量對照大全.xlsx
                    wb = load_workbook(r'C:\Users\林煥傑\Desktop\2020食物熱量對照大全.xlsx')
                    ws = wb.active
                    ws.append([food, float(cal)])
                    wb.save("C:\\Users\\林煥傑\\Desktop\\2020食物熱量對照大全.xlsx")


                    """
                    從基本資料更新當日可攝取熱量
                    """
                    infor = User_Info.objects.get(uid=uid,data_type="個人資料",number=1)

                    # 替使用者建立當日資料
                    if User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).exists() == False:
                        User_Info.objects.create(uid=uid,data_type="個人資料",name=name,user_bmr=infor.user_bmr,cal=infor.user_bmr)

                    user = User_Info.objects.get(uid=uid,data_type="個人資料",date=today)  # 當天的使用者紀錄
                    calculate = float(user.cal) - float(cal)  # 計算當日熱量並更新當日紀錄
                    User_Info.objects.filter(uid=uid,data_type="個人資料",date=today).update(cal=str(calculate))

                    message.append(TextSendMessage(text=food + '預估的熱量是' + cal + 'Kcal'))
                    message.append(TextSendMessage(text='今天仍可以攝取' + str(round(calculate, 3)) + 'Kcal'))

                elif '每月報告' in mtext:  #  5.每月報告，輸出折線圖

                    # 先取得使用者基本資料
                    infor = User_Info.objects.get(uid=uid,data_type="個人資料",number=1)

                    dateList = []
                    tdeeList = []
                    xList = []
                    month_dict = dict()
                    # date = pd.date_range(end=today, periods=31)
                    for i in range(31):
                        dateBefore = today - datetime.timedelta(days=30-i)
                        dateBeforechange = dateBefore.strftime('%m/%d')
                        dateList.append(dateBeforechange)
                        xList.append(i+1)
                        try:
                            user = User_Info.objects.get(uid=uid, data_type="個人資料",date=dateBefore)
                            user_tdee = float(user.user_bmr) - float(user.cal)
                            tdeeList.append(user_tdee)
                            month_dict[dateBeforechange] = user_tdee
                        except User_Info.DoesNotExist:
                                tdeeList.append(0)
                    tdeePlot = [float(infor.user_bmr)]*31

                    month_str = ''
                    over = 0
                    over_date = ''
                    for date in month_dict:
                        if month_dict[date] != 0.0:
                            month_str += date + ':' + str(int(month_dict[date])) + '大卡  '
                            if month_dict[date] > float(infor.user_bmr):
                                over += 1
                                over_date += date + ' '
                        
                    message.append(TextSendMessage(text=month_str))
                    message.append(TextSendMessage(text='有' + str(over) +'天超過標準，日期是： ' + over_date))

                    py.clf()
                    py.plot(xList, tdeeList, label = "Consumption Level", marker = 'o')
                    py.plot(xList, tdeePlot, label = "TDEE Level", marker = 'o')
                    py.legend(loc = 'upper left')
                    py.xlabel('Dates')
                    py.ylabel('Tdee')
                    py.title('Yours Past 30 Day\'s Consumption Record')
                    py.xticks(xList)
                    plotName = str(random.random())

                    # 取得折線圖URL
                    py.savefig('C:\\Users\\林煥傑\\Desktop\\' + plotName + '.png')
                    CLIENT_ID = "45c5f1f50f87a64"
                    PATH = "C:\\Users\\林煥傑\\Desktop\\" + plotName +  ".png"
                    im = pyimgur.Imgur(CLIENT_ID)
                    uploaded_image = im.upload_image(PATH, title="Uploaded with PyImgur")

                    # 輸出圖表
                    message.append(ImageSendMessage(original_content_url=uploaded_image.link, preview_image_url=uploaded_image.link))

                elif "." in mtext:  # 6.更改飲食策略
                    strategy = mtext[:1]
                    user = User_Info.objects.get(uid=uid, data_type='個人資料', number=1)
                    if strategy == 'a' :
                        message.append(TextSendMessage(text='在增重的策略下，你每天要攝取超過' + str(round(float(user.user_bmr), 3)) + ' Kcal/Day。'))
                    elif strategy == 'b':
                        message.append(TextSendMessage(text='在維持體重的情況下，你/妳每天可以攝取的熱量是' + str(round(float(user.user_bmr), 3)) + ' Kcal/Day。'))
                    elif strategy == 'c':
                        message.append(TextSendMessage(text='在減重的策略下，你不能攝取超過' + str(round(float(user.user_bmr), 3)) + ' Kcal/Day。'))
                    else:
                        message.append(TextSendMessage(text='輸入格式為 英文小寫加策略 \n(a:增重, b:正常, c:減重)'))

                elif '更改策略' in mtext:  # 更改飲食策略(快捷鍵)
                    message.append(TextSendMessage(text="要改成什麼策略？",
                            quick_reply=QuickReply(
                                items=[
                                    QuickReplyButton(
                                        action=MessageAction(label="增重",text="a.增重")
                                    ),
                                    QuickReplyButton(
                                        action=MessageAction(label="正常",text="b.正常")
                                    ),
                                    QuickReplyButton(
                                        action=MessageAction(label="減重",text="c.減重")
                                    )
                                ]
                            )
                        )
                    )
                elif "使用說明" in mtext:  # 7.使用說明
                    message.append(TextSendMessage(text="希望這個小工具能幫助你走上\n健康快樂的飲食之路\n" + "以下為使用說明手冊\n1.在開始使用功能之前，輸入或修改基本資料：\n輸入格式為m/f(男/女) 體重 身高 年齡 運動習慣”，以上資訊皆須按照順序且以半形空格隔開\n舉例：m 18 185 80 C\n2.運動習慣有五個選項，請選取以下五個英文字母其中一個輸入：\nA：久坐\nB：輕量活動(1~3天/週)\nC：中度活動量(3~5天/週)\nD：高度活動量(6~7天/週)\nE：非常高度活動量(運動員)\n3.輸入食物的方法為“用餐時間/食物名稱”\n用餐時間：早餐、午餐、晚餐、宵夜、飲料、其他\n舉例：輸入“早餐/蛋餅”\n4.自行輸入食物熱量的方式為“時段，食物，熱量”，以上資訊皆須按照順序且以全形逗號隔開\n舉例：輸入“早餐，蛋餅，100”\n5.查詢選單內容包含：使用說明、查詢30日資料、查詢個人資料（一天還能吃多少）及修改策略"))

                else:  # 提醒部分，我們沒做出主動提醒，所以改成使用者輸入無法啟動功能的訊息就傳提醒訊息
                    try:   
                        line_bot_api.push_message(uid, TextSendMessage(text='你今天記錄熱量了嗎？'))
                    except LineBotApiError as e:
                        #error handle
                        raise e

                line_bot_api.reply_message(event.reply_token,message)

        return HttpResponse()
    else:
        return HttpResponseBadRequest()
